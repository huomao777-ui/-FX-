#!/usr/bin/env python3
"""
新闻模板三边同步脚本：news_templates.json ↔ news_library_*.xlsx ↔ news_library_*.csv
用法：
  python sync_news_data.py         # 从 JSON 同步到 XLSX 和 CSV
  python sync_news_data.py --check # 只检查一致性，不写入
"""

import json, csv, os, sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(BASE_DIR, "news_templates.json")
CSV_WITH_IMAGES = os.path.join(BASE_DIR, "news_library_with_images.csv")
CSV_TEXT_ONLY = os.path.join(BASE_DIR, "news_library_text_only.csv")
XLSX_WITH_IMAGES = os.path.join(BASE_DIR, "news_library_with_images.xlsx")
XLSX_TEXT_ONLY = os.path.join(BASE_DIR, "news_library_text_only.xlsx")

# 所有字段（含图片编号列，有图/无图共用绝大部分字段）
COMMON_FIELDS = [
    "id", "分类", "有图",
    "图片编号",    # 有图时填写，无图时留空
    "适用地区",
    "标题模板", "正文",
    "现象总结", "趋势预测", "正文末尾提示",
    "默认方向", "可选方向", "默认强度", "可选强度",
    "反向逻辑说明",
    "底层接口参数",
    "备注",
]


def load_json():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def template_to_row(t, has_image):
    """将单个模板转换为 CSV/XLSX 行（字典）。"""
    regions = "|".join(t.get("regions", []))
    image_ids = "|".join(t.get("image_ids", []))
    params = "|".join(t.get("parameter_hints", []))
    tags = t.get("tags", [])
    tags_str = "|".join(tags) if isinstance(tags, list) else str(tags)

    # 构造完整的 body（同 CSV 当前逻辑）
    summary = t.get("summary_template", "")
    outlook = t.get("trend_outlook", "")
    tail = t.get("analysis_tail", "")
    body = f"{summary} {outlook} {tail}"

    return {
        "id": t["id"],
        "分类": t.get("category", ""),
        "有图": "是" if has_image else "否",
        "图片编号": image_ids if has_image else "",
        "适用地区": regions,
        "标题模板": t.get("headline_template", ""),
        "正文": body,
        "现象总结": summary,
        "趋势预测": outlook,
        "正文末尾提示": tail,
        "默认方向": t.get("default_bias", ""),
        "可选方向": "|".join(t.get("bias_options", [])),
        "默认强度": t.get("default_strength", ""),
        "可选强度": "|".join(t.get("strength_options", [])),
        "反向逻辑说明": t.get("reverse_logic_note", ""),
        "底层接口参数": params,
        "备注": tags_str,
    }


def write_csv(rows, csv_path):
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COMMON_FIELDS)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    print(f"  ✓ CSV: {csv_path}")


def write_xlsx(rows, xlsx_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(COMMON_FIELDS)
    for r in rows:
        ws.append([r.get(f, "") for f in COMMON_FIELDS])
    wb.save(xlsx_path)
    print(f"  ✓ XLSX: {xlsx_path}")


def sync():
    data = load_json()
    templates = data.get("templates", [])
    print(f"共 {len(templates)} 个模板")

    with_images = []
    text_only = []

    for t in templates:
        if t.get("has_image", False):
            with_images.append(template_to_row(t, True))
        else:
            text_only.append(template_to_row(t, False))

    print(f"  有图: {len(with_images)} 条")
    print(f"  无图: {len(text_only)} 条")

    # 写 CSV
    write_csv(with_images, CSV_WITH_IMAGES)
    write_csv(text_only, CSV_TEXT_ONLY)

    # 写 XLSX
    try:
        write_xlsx(with_images, XLSX_WITH_IMAGES)
        write_xlsx(text_only, XLSX_TEXT_ONLY)
    except Exception as e:
        print(f"  ! XLSX 写入失败: {e}")
        print("  (CSV 已写入，XLSX 请手动关闭后再运行)")


def check():
    """只检查，不写入"""
    from openpyxl import load_workbook

    def load_xlsx(path):
        wb = load_workbook(path)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            rows.append(str(row[0]))  # id 列
        return set(rows)

    data = load_json()
    templates = data.get("templates", [])
    json_ids = {t["id"] for t in templates}

    print("=== JSON ids ===")
    print(f"  JSON: {len(json_ids)} 条")
    for fid in sorted(json_ids):
        print(f"    {fid}")

    print(f"\n=== 有图 CSV ===")
    with open(CSV_WITH_IMAGES, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        csv_ids = {r["id"] for r in reader}
    print(f"  有图: {len(csv_ids)} 条")
    for fid in sorted(csv_ids):
        print(f"    {fid}")

    print(f"\n=== 无图 CSV ===")
    with open(CSV_TEXT_ONLY, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        csv_ids2 = {r["id"] for r in reader}
    print(f"  无图: {len(csv_ids2)} 条")
    for fid in sorted(csv_ids2):
        print(f"    {fid}")

    # cross check
    in_json_not_csv = json_ids - (csv_ids | csv_ids2)
    if in_json_not_csv:
        print(f"\n⚠️ JSON 有但 CSV 没有: {in_json_not_csv}")

    in_csv_not_json = (csv_ids | csv_ids2) - json_ids
    if in_csv_not_json:
        print(f"\n⚠️ CSV 有但 JSON 没有: {in_csv_not_json}")

    print(f"\n=== XLSX 验证 ===")
    # 按有图/无图分组分别比对
    json_with = {t["id"] for t in templates if t.get("has_image")}
    json_without = {t["id"] for t in templates if not t.get("has_image")}

    if os.path.exists(XLSX_WITH_IMAGES):
        xids = load_xlsx(XLSX_WITH_IMAGES)
        match = xids == json_with
        print(f"  {os.path.basename(XLSX_WITH_IMAGES)}: {len(xids)} 条 {'✓ 一致' if match else '✗ 不一致'}")
        if not match:
            print(f"    JSON有但XLSX无: {json_with - xids or '✓ 无'}")
            print(f"    XLSX有但JSON无: {xids - json_with or '✓ 无'}")

    if os.path.exists(XLSX_TEXT_ONLY):
        xids = load_xlsx(XLSX_TEXT_ONLY)
        match = xids == json_without
        print(f"  {os.path.basename(XLSX_TEXT_ONLY)}: {len(xids)} 条 {'✓ 一致' if match else '✗ 不一致'}")
        if not match:
            print(f"    JSON有但XLSX无: {json_without - xids or '✓ 无'}")
            print(f"    XLSX有但JSON无: {xids - json_without or '✓ 无'}")
    print()

    diff = json_ids ^ (csv_ids | csv_ids2)
    if diff:
        print(f"\n⚠️ JSON 和 CSV 不一致! 差异: {diff}")
        return False
    else:
        print(f"\n✓ JSON 和 CSV 完全一致")
        return True


if __name__ == "__main__":
    if "--check" in sys.argv:
        check()
    else:
        sync()
        print("\n✓ 同步完成")
